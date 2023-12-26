import openpyxl
def open_and_read_excel_file(sheetname, column_value, file):
    filename = file
    wb_obj = openpyxl.load_workbook(filename, data_only = True)
    sheet_obj = wb_obj.get_sheet_by_name(sheetname)
    column = int(column_value)
    m_row = sheet_obj.max_row
    my_list = [] #created an empty list
    for i in range(2, m_row+1): # Here I have started the loop from 2 as I want to skip the column heading value in output
        cell_obj = sheet_obj.cell(row=i, column=column)
        print(cell_obj.value)
        my_list.append(cell_obj.value)
    return my_list